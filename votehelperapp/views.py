from django.http import HttpResponse
from django.shortcuts import render, redirect
from .forms import VoterForm, CreateForm
from .models import Voter
from django.views.generic import ListView
import csv
from django.contrib.auth import authenticate, login
from django.contrib.auth.forms import UserCreationForm
from django.contrib.auth import get_user_model
from django.contrib.auth.models import Group, User
from django.contrib.auth.decorators import login_required
import openpyxl, io
from django.contrib.auth.mixins import LoginRequiredMixin

class VoterList(LoginRequiredMixin, ListView):
    template_name = "voterlist.html"
    model = Voter

def loginUser(request):
    if request.method == 'POST':
        username = request.POST['username']
        password = request.POST['password']
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            return redirect('voterlist')
        else:
            return render(request, 'login.html', {'error': 'Invalid username or password'})
    else:
        return render(request, 'login.html')

def register(request):
    if request.method == 'POST':
        form = UserCreationForm(request.POST)
        if form.is_valid():
            form.save()
            username = form.cleaned_data.get('username')
            raw_password = form.cleaned_data.get('password1')
            user = authenticate(username=username, password=raw_password)
            login(request, user)
            return redirect('stats')
    else:
        form = UserCreationForm()
    return render(request, 'register.html', {'form': form})

@login_required(login_url='/login')
def survey(request):
    form = VoterForm(request.POST or None, request.FILES or None)
    voterId = request.GET.get('id')  
    if voterId != None:
        voter = Voter.objects.get(id=voterId)
        form = VoterForm(instance=voter)
    if request.method == "POST":
        if voterId != None:
            form = VoterForm(request.POST, instance=voter)
        else:
            form = VoterForm(request.POST, request.FILES,)
        if form.is_valid():
            form.save()
            return redirect("/")
        else:
            print(form.errors)

    a = Voter.objects.filter(decision=1).count

    context = {'form': form, 'id':voterId, 'voter':voter}
    return render(request, "survey.html", context)

@login_required(login_url='/login')
def stats(request):
    # if request.user.is_superuser:
    #     return redirect('assign')

    yes = Voter.objects.filter(decision=1).count
    no = Voter.objects.filter(decision=2).count

    context = {'yes':yes, 'no':no}
    return render(request, "stats.html", context)

@login_required(login_url='/login')
def assign(request):
    user_pk = request.POST.get('user')
    group_name = request.POST.get('group')
    groups = Group.objects.all()
    User = get_user_model()
    users = User.objects.all()

    if request.method == 'POST':
        if user_pk and group_name:
            # Get the user and group objects
            user = User.objects.get(pk=user_pk)
            group = Group.objects.get(name=group_name)

            # Remove the user from their current group (if they are in one)
            if user.groups:
                user.groups.clear()

            # Add the user to the group
            group.user_set.add(user)

    context = {'users':users, 'groups':groups}
    return render(request, "assign.html", context)

@login_required(login_url='/login')
def assignIndividual(request):
    user_pk = request.POST.get('user')
    voter_names = request.POST.getlist('options')
    voters = Voter.objects.all()
    User = get_user_model()
    users = User.objects.all()

    if request.method == 'POST':
        # Get the user and group objects
        if user_pk.isdigit():
            user = User.objects.get(pk=user_pk)
            for voter_name in voter_names:
                voter = Voter.objects.get(name=voter_name)
                voter.assignedEmp = user.username
                voter.save()

    context = {'users':users, 'voters':voters}
    return render(request, "assignIndividual.html", context)

# @login_required
# @require_POST
@login_required(login_url='/login')
def import_csv(request):
    if request.method == 'POST':
        try:
            csv_file = request.FILES['csv_file']
            data = csv_file.read().decode("utf-8")
            if not csv_file.name.endswith(".csv"):
                return redirect("/")
            
            lines = data.split("\n")
            lines.pop(0)
            for line in lines:
                if not line:
                    continue

                reader = csv.reader([line])
                fields = next(reader)
                groupNum = fields[4].strip()

                # create the group
                group = Group.objects.filter(name=groupNum)

                if not group.exists():
                    group = Group.objects.create(name=groupNum)
                    
                voter = Voter(name=fields[0], address=fields[1], phone=fields[2], email=fields[3], neighbourhood=groupNum)
                voter.save()

        except Exception as e:
            print(e)
    return redirect("/")

@login_required(login_url='/login')
def addVoter(request):
    form = CreateForm(request.POST or None)
    if request.method == "POST":
        form = CreateForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect("/")
        else:
            print(form.errors)

    context = {'form': form}
    return render(request, "addvoter.html", context)

@login_required(login_url='/login')
def export(request):

    # Open a new Excel workbook
    workbook = openpyxl.Workbook()

    # Get the active worksheet
    worksheet = workbook.active

    # Set the worksheet title
    worksheet.title = "Voters"

    # Set the column headers
    worksheet.cell(row=1, column=1).value = "Name"
    worksheet.cell(row=1, column=2).value = "Address"
    worksheet.cell(row=1, column=3).value = "Phone"
    worksheet.cell(row=1, column=4).value = "Email"
    worksheet.cell(row=1, column=5).value = "Decision"
    worksheet.cell(row=1, column=6).value = "Notes"

    # Get the voters from the database
    voters = Voter.objects.all()

    # Loop through the voters and add them to the worksheet
    for i, voter in enumerate(voters):
        worksheet.cell(row=i+2, column=1).value = voter.name
        worksheet.cell(row=i+2, column=2).value = voter.address
        worksheet.cell(row=i+2, column=3).value = voter.phone
        worksheet.cell(row=i+2, column=4).value = voter.email
        if voter.decision == "1":
            worksheet.cell(row=i+2, column=5).value = "Support"
        elif voter.decision == "2":
            worksheet.cell(row=i+2, column=5).value = "Not Supporting"
        elif voter.decision == "3":
            worksheet.cell(row=i+2, column=5).value = "Undecided"
        elif voter.decision == "4":
            worksheet.cell(row=i+2, column=5).value = "Not Available"
        worksheet.cell(row=i+2, column=6).value = voter.notes

    # Set the column widths to fit the cell contents
    worksheet.column_dimensions['A'].width = 20
    worksheet.column_dimensions['B'].width = 50
    worksheet.column_dimensions['C'].width = 20
    worksheet.column_dimensions['D'].width = 30
    worksheet.column_dimensions['E'].width = 20
    worksheet.column_dimensions['F'].width = 50

    # Save the workbook to a memory buffer
    buffer = io.BytesIO()
    workbook.save(buffer)

    response = HttpResponse(buffer.getvalue())
    response['Content-Disposition'] = 'attachment; filename=voters.xlsx'
    response['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    return response

